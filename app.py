from flask import Flask, request, jsonify
from flask_cors import CORS
import cv2
import numpy as np
from deepface import DeepFace
from collections import Counter
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
CORS(app)

# Load OpenCV's built-in face detector
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Function to categorize age into age groups
def get_age_group(age):
    if age <= 12:
        return "Child (0-12)"
    elif age <= 19:
        return "Teenager (13-19)"
    elif age <= 34:
        return "Young Adult (20-34)"
    elif age <= 49:
        return "Middle-aged Adult (35-49)"
    else:
        return "Senior (50+)"

# Function to log detection results into an Excel file
def log_detection_to_excel(age_group, gender):
    file_name = "ad_stats_log.xlsx"

    if not os.path.exists(file_name):
        # Create a new workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Age Group", "Gender"])
        wb.save(file_name)

    # Load existing workbook
    wb = load_workbook(file_name)
    ws = wb.activex

    # Append a new row with current detection info
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([timestamp, age_group, gender])
    wb.save(file_name)

# Health check route
@app.route("/", methods=["GET"])
def home():
    return jsonify({"message": "Server is running!"}), 200

# Route for predicting age and gender
@app.route("/predict-age-gender", methods=["POST"])
def predict_age_gender():
    if "image" not in request.files:
        return jsonify({"error": "No image file found"}), 400

    file = request.files["image"]
    img = cv2.imdecode(np.frombuffer(file.read(), np.uint8), cv2.IMREAD_COLOR)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Detect faces using OpenCV
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

    if len(faces) == 0:
        return jsonify({"error": "No faces detected"}), 400

    print(f"Detected {len(faces)} faces")

    age_groups = []
    genders = []

    for (x, y, w, h) in faces:
        face = img[y:y+h, x:x+w]

        try:
            result = DeepFace.analyze(face, actions=["age", "gender"], enforce_detection=False)[0]
            age = result.get("age", 0)
            gender_data = result.get("gender", "")
            gender = max(gender_data, key=gender_data.get) if isinstance(gender_data, dict) else str(gender_data)

            age_group = get_age_group(age)
            print(f"Face detected - Age: {age}, Gender: {gender}")
            print(f"Age group: {age_group}")

            age_groups.append(age_group)
            genders.append(gender)
        except Exception as e:
            print(f"Skipping one face due to error: {str(e)}")
            continue

    if not age_groups or not genders:
        return jsonify({"error": "Analysis failed for all detected faces"}), 500

    # Get majority values
    majority_age_group = Counter(age_groups).most_common(1)[0][0]
    majority_gender = Counter(genders).most_common(1)[0][0]

    # Log to Excel
    log_detection_to_excel(majority_age_group, majority_gender)

    return jsonify({
        "age_group": majority_age_group,
        "gender": majority_gender
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
